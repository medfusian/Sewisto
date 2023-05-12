import psycopg2.extras
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from io import BytesIO
from flask import render_template, redirect, request, url_for, session, g, flash, abort, send_file, Response
from . import cart_bp


@cart_bp.context_processor
def inject_user():
    current_user = None
    user = None
    admin = False

    if 'user_id' in session:
        g.cursor.execute("SELECT * FROM public.user WHERE id = %s", (session['user_id'],))
        current_user = g.cursor.fetchone()
        user = current_user[3]
        current_user = current_user[1]
        admin = False
    elif 'admin_id' in session:
        admin = True

    return dict(current_user=current_user, user=user, admin=admin)


@cart_bp.route('/cart')
def cart():
    # Проверяем, что пользователь аутентифицирован
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    g.cursor.execute("SELECT * FROM cart WHERE user_id=%s AND order_id IS NULL", (session['user_id'],))
    cart = g.cursor.fetchall()

    total_sum = 0
    for c in cart:
        total_sum += c[5]

    return render_template('cart.html', cart=cart, total_sum=total_sum, user_id=session['user_id'])


@cart_bp.route('/add_to_cart/<int:product_id>', methods=['POST'])
def add_to_cart(product_id):
    # Проверяем, что пользователь аутентифицирован
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    g.cursor.execute("SELECT * FROM product WHERE article=%s", (product_id,))
    product = g.cursor.fetchone()
    if not product:
        flash('Товар не найден', 'error')
        return redirect(url_for('main.index'))

    g.cursor.execute("SELECT * FROM cart")
    cart = g.cursor.fetchall()
    is_on_cart = False

    for c in cart:
        if c[1] == product[0]:
            if c[6] == session['user_id'] and c[7] is None:
                is_on_cart = True
                cart_id = c[0]

    if is_on_cart:
        g.cursor.execute("UPDATE cart SET count=count+1 WHERE id=%s AND user_id=%s", (cart_id, session['user_id'],))
    else:
        g.cursor.execute("INSERT INTO cart (article, name, price, count, user_id) VALUES (%s, %s, %s, %s, %s)",
                         (product[0], product[1], product[7], 1, session['user_id']))

    flash(f'Товар "{product[1]}" добавлен в корзину', 'success')
    return redirect(url_for('cart.cart'))


@cart_bp.route('/edit_cart/<int:cart_id>', methods=['GET', 'POST'])
def update_count(cart_id):
    # Получаем товар по его идентификатору
    with g.connect.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
        cursor.execute('SELECT * FROM cart WHERE id = %s', (cart_id,))
        cart = cursor.fetchone()

    # Если товар не найден, возвращаем ошибку 404
    if not cart:
        abort(404)

    # Обработка GET запроса
    if request.method == 'GET':
        # Отображаем форму для редактирования товара
        return render_template('edit_cart.html', product=cart)

    # Обработка POST запроса
    if request.method == 'POST':
        # Обновляем информацию о товаре в базе данных
        count = request.form['count']

    with g.connect.cursor() as cursor:
        cursor.execute(
            'UPDATE cart SET count=%s WHERE id=%s', (count, cart_id))

    # Перенаправляем пользователя на страницу с корзиной
    return redirect(url_for('cart.cart'))


@cart_bp.route('/delete_cart/<int:cart_id>', methods=['POST'])
def delete_product(cart_id):
    # Удаляем товар из таблицы product
    g.cursor.execute("DELETE FROM cart WHERE id=%s", (cart_id,))

    # Возвращаемся на страницу со списком товаров
    flash('Товар успешно удален', 'success')
    return redirect(url_for('cart.cart'))


@cart_bp.route('/create_order/<int:user_id>', methods=['POST', 'GET'])
def create_order(user_id):
    # Проверяем, что пользователь аутентифицирован
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    g.cursor.execute("INSERT INTO public.order (user_id) values (%s) RETURNING id", (user_id,))
    order_id = g.cursor.fetchone()

    g.cursor.execute("UPDATE cart SET order_id=%s WHERE order_id IS NULL AND user_id=%s", (order_id, user_id,))

    # Перенаправляем пользователя на страницу с корзиной
    return redirect(url_for('cart.cart'))


@cart_bp.route('/orders')
def orders():
    # Проверяем, что пользователь аутентифицирован
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    g.cursor.execute("SELECT o.id, o.order_date, SUM(c.summary) AS total_summary FROM cart c "
                     "JOIN public.order o on o.id = c.order_id "
                     "WHERE o.user_id=%s GROUP BY o.id", (session['user_id'],))
    orders = g.cursor.fetchall()

    total_sum = 0
    for order in orders:
        total_sum += order[2]

    return render_template('orders.html', orders=orders, total_sum=total_sum)


@cart_bp.route('/admin/orders/')
def admin_orders():
    # Проверяем, что пользователь аутентифицирован как администратор
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    g.cursor.execute("SELECT * FROM public.order")
    orders = g.cursor.fetchall()
    return render_template('admin/admin_orders.html', orders=orders)


@cart_bp.route('/cart/sort_price')
def sort_price():
    g.cursor.execute("SELECT * FROM cart WHERE user_id=%s AND order_id IS NULL ORDER BY price DESC",
                     (session['user_id'],))
    cart = g.cursor.fetchall()

    total_sum = 0
    for c in cart:
        total_sum += c[5]

    return render_template('cart.html', cart=cart, total_sum=total_sum, user_id=session['user_id'])


@cart_bp.route('/cart/sort_count')
def sort_count():
    g.cursor.execute("SELECT * FROM cart WHERE user_id=%s AND order_id IS NULL ORDER BY count DESC",
                     (session['user_id'],))
    cart = g.cursor.fetchall()

    total_sum = 0
    for c in cart:
        total_sum += c[5]

    return render_template('cart.html', cart=cart, total_sum=total_sum, user_id=session['user_id'])


@cart_bp.route('/admin/order_report')
def order_report():
    # Проверяем, что пользователь аутентифицирован как администратор
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Отчет по заказам'

    g.cursor.execute(
        "SELECT o.id, o.user_id, CAST(o.order_date as varchar(19)), SUM(c.summary) AS total_summary FROM cart c "
        "JOIN public.order o on o.id = c.order_id "
        "GROUP BY o.id")
    orders = g.cursor.fetchall()

    worksheet.cell(row=1, column=1, value='ID')
    worksheet.cell(row=1, column=2, value='Пользователь')
    worksheet.cell(row=1, column=3, value='Дата')
    worksheet.cell(row=1, column=4, value='Сумма')

    for row_idx, row_data in enumerate(orders, start=2):
        worksheet.cell(row=row_idx, column=1, value=row_data[0])
        worksheet.cell(row=row_idx, column=2, value=row_data[1])
        worksheet.cell(row=row_idx, column=3, value=row_data[2])
        worksheet.cell(row=row_idx, column=4, value=row_data[3])

    workbook.save('Отчет по заказам.xlsx')

    return redirect(url_for('admin.admin_panel'))


@cart_bp.route('/admin/order_analytics')
def order_analytics():
    # Проверяем, что пользователь аутентифицирован как администратор
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))

    # выполнить запрос к базе данных
    g.cursor.execute("SELECT to_char(order_date, 'YYYY - MM') AS month, COUNT(*) AS count "
                     "FROM public.order GROUP BY month;")
    results = g.cursor.fetchall()
    print(results)

    # разбить результаты на списки месяцев и количества заказов
    months = [r[0] for r in results]
    counts = [r[1] for r in results]

    print(months)
    print(counts)

    # построить график
    fig = Figure()
    axis = fig.add_subplot(1, 1, 1)
    axis.plot(months, counts)
    axis.set_xlabel('Month')
    axis.set_ylabel('Number of orders')
    axis.set_title('Orders by Month')
    axis.set_ylim([0, 10])  # ограничить ось y до 10 единиц

    # создать объект canvas для отображения графика
    canvas = FigureCanvas(fig)

    # создать объект для хранения изображения графика в памяти
    output = BytesIO()
    canvas.print_png(output)

    # создать объект Response с содержимым изображения и соответствующим заголовком
    response = Response(output.getvalue(), mimetype='image/png')
    response.headers['Content-Disposition'] = 'attachment; filename=orders_by_month.png'

    # вернуть объект Response, который будет отображен в новой вкладке браузера
    return response

